"""French cash reconciliation — everything *around* the Zelty feed.

The workbook "Suivi des especes 2026.xlsx" holds one sheet per restaurant. Each
sheet is a daily cash log with these columns (row 5 headers, data from row 6):

  B  Date
  C  Caisse            fond de caisse / till float (constant per site)
  D  Zelty             cash sales reported by Zelty  <-- fed by main.py / the API
  E  Coffre            running safe balance          (formula)
  F  A deposer         amount to bank                (formula)
  G  Depot             actual deposit banked         (manual)
  H  FC Espece         cash physically counted       (manual)
  I  FC Sortie Caisse  cash taken out of the till     (manual)
  J  Ecart de caisse   discrepancy = D - H           (formula)  <-- the recon result
  K  Vraiment Caisse   H - I                         (formula)
  L  Commentaires      free text                     (manual)
  M  Justificatif      free text                     (manual)

The formulas, taken straight from the sheet:
  E6 (first row)  = E2 (target safe float) + opening seed
  E(n)            = E(n-1) + D(n) - G(n-1)
  F(n)            = E(n) - E2 - I(n)
  J(n)            = D(n) - H(n)
  K(n)            = H(n) - I(n)

While we wait for the Zelty API key, D is read from the historical sheet. Once
the key arrives, `main.py` supplies D per (site, date) and the same engine runs.
The engine here is validated (see `validate`) to reproduce the sheet's own J/K
values exactly, so we know column D is the only missing input.
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# The workbook and this report use accented French text and the ± / − signs.
# The default Windows console codepage (cp1252) can't encode some of them, so
# force UTF-8 on our streams.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

WORKBOOK = Path(__file__).with_name("Suivi des espèces 2026.xlsx")

# Column letters -> field, for the daily log (headers on row 5, data from row 6).
COL = {
    "date": "B",
    "caisse": "C",
    "zelty": "D",
    "coffre": "E",
    "a_deposer": "F",
    "depot": "G",
    "fc_espece": "H",
    "fc_sortie": "I",
    "ecart": "J",
    "vraiment_caisse": "K",
    "commentaire": "L",
    "justificatif": "M",
}
HEADER_ROW = 5
FIRST_DATA_ROW = 6

# Sheet name (workbook) -> Zelty restaurant. Filled in once we see the live
# `/restaurants` response from main.py. Left as identity hints for now; the
# reconciler tolerates a missing entry (it just can't match that site's Zelty).
SITE_TO_ZELTY: dict[str, str] = {
    # "Bercy": "<zelty restaurant id or exact name>",
    # "VDE": "...",
}


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _numopt(v) -> float | None:
    """Like _num but keeps None (a formula cell with no cached result — e.g.
    right after an openpyxl save, before Excel recalculates)."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


@dataclass
class DayRow:
    site: str
    date: dt.date | None
    caisse: float = 0.0
    zelty: float = 0.0          # D — from the sheet today, from the API later
    depot: float = 0.0          # G
    fc_espece: float = 0.0      # H
    fc_sortie: float = 0.0      # I
    # stored (as-in-sheet) derived values, for validation. None = the formula
    # cell had no cached result (skip it when validating).
    coffre_sheet: float | None = 0.0
    a_deposer_sheet: float | None = 0.0
    ecart_sheet: float | None = 0.0
    vraiment_sheet: float | None = 0.0
    commentaire: str = ""
    justificatif: str = ""

    # computed by the engine
    coffre: float = 0.0
    a_deposer: float = 0.0
    ecart: float = 0.0
    vraiment_caisse: float = 0.0

    @property
    def has_activity(self) -> bool:
        return any([self.zelty, self.depot, self.fc_espece, self.fc_sortie])


@dataclass
class Site:
    name: str
    target_safe: float = 0.0    # E2 — safe float target used by column F
    opening_coffre: float = 0.0  # E6 as read from the sheet (first-row seed)
    rows: list[DayRow] = field(default_factory=list)


def _cell(ws, col_field: str, row: int):
    return ws[f"{COL[col_field]}{row}"].value


def load_workbook_sites(path: Path = WORKBOOK) -> list[Site]:
    """Read every restaurant sheet into a list of Site objects."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sites: list[Site] = []
    for ws in wb.worksheets:
        site = Site(
            name=ws.title,
            target_safe=_num(ws["E2"].value),
        )
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            raw_date = _cell(ws, "date", r)
            date = raw_date.date() if isinstance(raw_date, dt.datetime) else None
            if date is None and _cell(ws, "zelty", r) in (None, ""):
                continue  # blank trailing row
            row = DayRow(
                site=ws.title,
                date=date,
                caisse=_num(_cell(ws, "caisse", r)),
                zelty=_num(_cell(ws, "zelty", r)),
                depot=_num(_cell(ws, "depot", r)),
                fc_espece=_num(_cell(ws, "fc_espece", r)),
                fc_sortie=_num(_cell(ws, "fc_sortie", r)),
                coffre_sheet=_numopt(_cell(ws, "coffre", r)),
                a_deposer_sheet=_numopt(_cell(ws, "a_deposer", r)),
                ecart_sheet=_numopt(_cell(ws, "ecart", r)),
                vraiment_sheet=_numopt(_cell(ws, "vraiment_caisse", r)),
                commentaire=str(_cell(ws, "commentaire", r) or ""),
                justificatif=str(_cell(ws, "justificatif", r) or ""),
            )
            site.rows.append(row)
        if site.rows:
            site.opening_coffre = site.rows[0].coffre_sheet or 0.0
        sites.append(site)
    return sites


def compute(site: Site, zelty_by_date: dict[dt.date, float] | None = None) -> None:
    """Recompute the derived columns (E, F, J, K) for a site in place.

    If `zelty_by_date` is given, it overrides column D per date (this is how the
    live Zelty feed plugs in). Otherwise the sheet's own D values are used.

    Mirrors the sheet formulas exactly:
        E(n) = E(n-1) + D(n) - G(n-1)   [first row seeded from the sheet]
        F(n) = E(n) - target_safe - I(n)
        J(n) = D(n) - H(n)
        K(n) = H(n) - I(n)
    """
    prev_coffre = None
    prev_depot = 0.0
    for i, row in enumerate(site.rows):
        if zelty_by_date and row.date in zelty_by_date:
            row.zelty = zelty_by_date[row.date]

        if i == 0:
            row.coffre = site.opening_coffre
        else:
            row.coffre = prev_coffre + row.zelty - prev_depot

        row.a_deposer = row.coffre - site.target_safe - row.fc_sortie
        row.ecart = row.zelty - row.fc_espece
        row.vraiment_caisse = row.fc_espece - row.fc_sortie

        prev_coffre = row.coffre
        prev_depot = row.depot


def validate(sites: list[Site], tol: float = 0.01) -> tuple[list[str], int]:
    """Confirm the engine reproduces the sheet's stored J/K/F values from D/H/I.

    Returns (mismatch descriptions, number of cells actually compared). Cells
    whose cached formula result is missing are skipped, so `checked` can be 0
    right after an openpyxl save (Excel recalculates cached values on open).
    Coffre (E) is not checked row-by-row here because its first-row seed varies;
    the running recurrence is exercised via F which depends on E.
    """
    problems: list[str] = []
    checked = 0
    for site in sites:
        compute(site)  # uses the sheet's own D
        for row in site.rows:
            for label, got, want in (
                ("Ecart(J)", row.ecart, row.ecart_sheet),
                ("Vraiment(K)", row.vraiment_caisse, row.vraiment_sheet),
                ("Coffre(E)", row.coffre, row.coffre_sheet),
                ("ADeposer(F)", row.a_deposer, row.a_deposer_sheet),
            ):
                if want is None:
                    continue  # formula not yet recalculated by Excel — can't compare
                checked += 1
                if abs(got - want) > tol:
                    problems.append(
                        f"{site.name} {row.date} {label}: engine={got:.2f} sheet={want:.2f} "
                        f"(diff {got - want:+.2f})"
                    )
    return problems, checked


def reconciliation_report(sites: list[Site], flag_threshold: float = 1.0) -> None:
    """Print the per-site ecart report from whatever D values are loaded.

    A day is only a genuine discrepancy when the store actually counted its cash
    (H > 0). Days with Zelty sales but no count (H == 0) are "not counted" —
    missing data, not a real écart — and are reported separately so they don't
    swamp the discrepancies that need explaining.
    """
    print("\nCash reconciliation — Ecart de caisse (Zelty D  −  counted H)")
    print("=" * 72)
    grand_disc = grand_missing = 0
    for site in sites:
        compute(site)
        active = [r for r in site.rows if r.has_activity]
        discrepancies = [r for r in active
                         if r.zelty and r.fc_espece and abs(r.ecart) >= flag_threshold]
        not_counted = [r for r in active if not r.fc_espece and r.zelty]
        zelty_missing = [r for r in active if not r.zelty and r.fc_espece]
        net = sum(r.ecart for r in discrepancies)
        print(f"\n  {site.name}   net écart {net:+.2f} over {len(discrepancies)} "
              f"discrepancy day(s); {len(not_counted)} not counted; "
              f"{len(zelty_missing)} awaiting Zelty")
        for r in discrepancies:
            note = r.commentaire or r.justificatif
            note = f"   — {note}" if note else ""
            print(f"    {r.date}  Zelty {r.zelty:8.2f}  counted {r.fc_espece:8.2f}  "
                  f"écart {r.ecart:+8.2f}{note}")
        grand_disc += len(discrepancies)
        grand_missing += len(not_counted) + len(zelty_missing)
    print("\n" + "-" * 72)
    print(f"  {grand_disc} genuine discrepancy day(s) and {grand_missing} "
          f"day(s) with data missing (not counted / awaiting Zelty) across "
          f"{len(sites)} sites")


def _last_data_row(ws):
    """Return the last row index whose date cell (col B) is populated, or None."""
    last = None
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws[f"{COL['date']}{r}"].value is not None:
            last = r
    return last


def site_last_state(site_name: str, path: Path = WORKBOOK) -> dict:
    """Last recorded till float + date for a site, for the UI to pre-fill.

    Returns {"last_date": date|None, "last_caisse": float}. The till float is
    the "last put-in value" the €150-floor rule is applied to.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if site_name not in wb.sheetnames:
        raise KeyError(site_name)
    ws = wb[site_name]
    last = _last_data_row(ws)
    if last is None:
        return {"last_date": None, "last_caisse": 0.0}
    raw_date = ws[f"{COL['date']}{last}"].value
    return {
        "last_date": raw_date.date() if isinstance(raw_date, dt.datetime) else None,
        "last_caisse": _num(ws[f"{COL['caisse']}{last}"].value),
    }


def day_values(site_name: str, date: dt.date, path: Path = WORKBOOK) -> dict:
    """Read the row already recorded for `date`, if any.

    Returns {"found": bool, "caisse", "zelty", "counted", "sortie", "depot",
    "gap"} where gap = Zelty(D) − counted(H). This is what lets the UI show the
    reconciliation that is *already in the sheet* for a historical date, rather
    than inventing mock numbers. gap is recomputed from D/H (not read from the J
    formula cell, whose cached value may be absent after an openpyxl save)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if site_name not in wb.sheetnames:
        raise KeyError(site_name)
    ws = wb[site_name]
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        v = ws[f"{COL['date']}{r}"].value
        d = v.date() if isinstance(v, dt.datetime) else v
        if d == date:
            zelty = _num(ws[f"{COL['zelty']}{r}"].value)
            counted = _num(ws[f"{COL['fc_espece']}{r}"].value)
            return {
                "found": True,
                "caisse": _num(ws[f"{COL['caisse']}{r}"].value),
                "zelty": zelty,
                "counted": counted,
                "sortie": _num(ws[f"{COL['fc_sortie']}{r}"].value),
                "depot": _num(ws[f"{COL['depot']}{r}"].value),
                "gap": zelty - counted,
            }
    return {"found": False}


def append_day(
    site_name: str,
    date: dt.date,
    zelty: float | None = None,
    counted: float = 0.0,      # H — FC Espèce (physically counted cash)
    sortie: float = 0.0,       # I — FC Sortie Caisse
    depot: float = 0.0,        # G — Dépôt banked
    comment: str = "",         # L — Commentaires (gap explanation)
    min_float: float = 150.0,  # €150 legal floor on the till float
    path: Path = WORKBOOK,
) -> dict:
    """Append one new daily row to a site's sheet, mirroring the row above.

    Business rules applied here:
      * Caisse (till float) = max(min_float, last recorded till float).
      * E/F/J/K are written as live formulas so Excel keeps recomputing them.
      * D (Zelty), H (counted), I (sortie), G (dépôt), L (comment) come from args.

    Cell styles/number formats are copied from the previous row. A `.bak` copy of
    the workbook is written first. Refuses to overwrite an existing date. Returns
    a summary dict (site, date, row, caisse, zelty, counted, gap, written).
    """
    import copy as _copy
    import shutil

    wb = openpyxl.load_workbook(path, data_only=False)
    if site_name not in wb.sheetnames:
        raise KeyError(f"No sheet named {site_name!r}. Sheets: {wb.sheetnames}")
    ws = wb[site_name]

    last = _last_data_row(ws)
    if last is None:
        raise ValueError(f"{site_name}: no existing data rows to extend from.")

    for r in range(FIRST_DATA_ROW, last + 1):
        v = ws[f"{COL['date']}{r}"].value
        d = v.date() if isinstance(v, dt.datetime) else v
        if d == date:
            return {"site": site_name, "date": date.isoformat(), "row": r,
                    "written": False,
                    "message": f"{date} already present on row {r}; skipped."}

    new = last + 1
    last_caisse = _num(ws[f"{COL['caisse']}{last}"].value)
    caisse = max(min_float, last_caisse)          # "last value, minimum €150"
    zelty_val = float(zelty) if zelty is not None else 0.0
    gap = zelty_val - float(counted)              # J = D - H

    values = {
        "date": dt.datetime(date.year, date.month, date.day),
        "caisse": caisse,
        "zelty": zelty_val,
        "coffre": f"=+E{last}+D{new}-G{last}",
        "a_deposer": f"=+E{new}-$E$2-I{new}",
        "depot": float(depot),
        "fc_espece": float(counted),
        "fc_sortie": float(sortie),
        "ecart": f"=D{new}-H{new}",
        "vraiment_caisse": f"=H{new}-I{new}",
        "commentaire": comment or None,
    }
    for field_name, value in values.items():
        if field_name not in COL:
            continue
        col = COL[field_name]
        src = ws[f"{col}{last}"]
        dst = ws[f"{col}{new}"]
        dst.value = value
        dst.number_format = src.number_format
        if src.has_style:
            dst.font = _copy.copy(src.font)
            dst.border = _copy.copy(src.border)
            dst.fill = _copy.copy(src.fill)
            dst.alignment = _copy.copy(src.alignment)

    # Ask Excel to recalculate (and re-cache) all formulas when it next opens the
    # file, since saving here clears cached formula results.
    try:
        wb.calculation.fullCalcOnLoad = True
    except AttributeError:
        pass

    backup = path.with_suffix(path.suffix + ".bak")
    shutil.copyfile(path, backup)
    wb.save(path)
    return {"site": site_name, "date": date.isoformat(), "row": new,
            "caisse": round(caisse, 2), "zelty": round(zelty_val, 2),
            "counted": round(float(counted), 2), "gap": round(gap, 2),
            "written": True, "backup": backup.name}


def main() -> None:
    ap = argparse.ArgumentParser(description="French cash reconciliation engine.")
    ap.add_argument("--validate", action="store_true",
                    help="Check the engine reproduces the sheet's own J/K/E/F.")
    ap.add_argument("--threshold", type=float, default=1.0,
                    help="Flag days whose |écart| is at least this (default 1.00).")
    ap.add_argument("--add-day", metavar="SITE",
                    help="Append a new daily row to SITE's sheet.")
    ap.add_argument("--date", help="Date for --add-day (YYYY-MM-DD; default today).")
    ap.add_argument("--zelty", type=float,
                    help="Zelty cash value for --add-day (default 0 / blank).")
    args = ap.parse_args()

    if args.add_day:
        date = (dt.date.fromisoformat(args.date) if args.date else dt.date.today())
        res = append_day(args.add_day, date, zelty=args.zelty)
        if res["written"]:
            print(f"{res['site']}: added {res['date']} on row {res['row']} "
                  f"(Caisse={res['caisse']}, Zelty={res['zelty']}, "
                  f"gap={res['gap']:+.2f}). Backup: {res['backup']}")
        else:
            print(f"{res['site']}: {res['message']}")
        return

    sites = load_workbook_sites()

    if args.validate:
        problems, checked = validate(sites)
        if checked == 0:
            print("Nothing to validate: no cached formula results in the "
                  "workbook.\nOpen it in Excel once (which recalculates and "
                  "re-caches formulas), then re-run --validate.")
        elif not problems:
            print(f"OK — engine reproduces J/K/E/F on all {checked} compared "
                  f"cells across {len(sites)} sites.")
        else:
            print(f"{len(problems)} mismatch(es) out of {checked} compared:")
            for p in problems[:50]:
                print("  " + p)
        return

    reconciliation_report(sites, flag_threshold=args.threshold)


if __name__ == "__main__":
    main()
